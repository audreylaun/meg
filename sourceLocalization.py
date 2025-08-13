import mne
import numpy as np
import argparse
import pandas as pd
import matplotlib.pyplot as plt
from mne.minimum_norm import apply_inverse, make_inverse_operator
import mne
import openpyxl
import os

def source_localization(sub):

    # Setting parameters to be used by both production and comprehension
    subjects_dir = '/Applications/freesurfer/7.4.1/subjects' # freesurfer MRI directory
    directory = '/Users/admin/Box Sync/Starling/Experiment1/MEG_data/' + sub + '/'

    conductivity = (0.3,) # single layer conductivity
    baseline_start = -300 #in milliseconds
    baseline_end = 0
    snr = 3.0
    method = "dSPM"
    parc = "aparc"
    loose = dict(surface=0.2, volume=1.0)

    # LOAD COMPREHENSION AND PRODUCTION DATA
    # load epochs
    epoch_fname_comp = directory + sub + '_comp_TEST-epo.fif'
    epochs_comp = mne.read_epochs(epoch_fname_comp)
    epoch_fname_prod = directory + sub + '_prod_TEST-epo.fif'
    epochs_prod = mne.read_epochs(epoch_fname_prod)

    # Get initial number of epochs for each condition
    prod_initial = len(epochs_prod)
    comp_initial = len(epochs_comp)

    # Reject bad epochs that have a max peak to peak signal amplitude that exceeds 3 picoteslas
    reject_criteria = dict(mag=3000e-15)
    epochs_prod.drop_bad(reject=reject_criteria)
    # print(epochs_prod.drop_log)
    epochs_comp.drop_bad(reject=reject_criteria)
    # print(epochs_comp.drop_log)

    # Calculate the number of epochs dropped
    prod_dropped = prod_initial - len(epochs_prod)
    comp_dropped = comp_initial - len(epochs_comp)
    # Save number of dropped epochs to excel sheet
    excel_path = "/Users/admin/Box Sync/Starling/Experiment1/epoch_dropping.xlsx"
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    if sub in header:
        col_idx = header.index(sub) + 1  # openpyxl is 1-indexed
    else:
        col_idx = len(header) + 1
        ws.cell(row=1, column=col_idx, value=sub)
    ws.cell(row=2, column=col_idx, value=prod_dropped)
    ws.cell(row=3, column=col_idx, value=comp_dropped)
    wb.save(excel_path)

    # Apply empty room SSPs, obtained during preprocessing
    epochs_prod.apply_proj()
    epochs_comp.apply_proj()

    # Equalize epoch counts
    ident_prod = epochs_prod['production identical']
    unrel_prod = epochs_prod['production unrelated']
    ident_comp = epochs_comp['comprehension identical']
    unrel_comp = epochs_comp['comprehension unrelated']
    mne.epochs.equalize_epoch_counts([ident_prod, ident_comp, unrel_prod, unrel_comp], method="random")

    # Create evoked objects low pass filtered at 40 Hz
    ident_prod = ident_prod.average().filter(l_freq=0, h_freq=40)
    unrel_prod = unrel_prod.average().filter(l_freq=0, h_freq=40)
    ident_comp = ident_comp.average().filter(l_freq=0, h_freq=40)
    unrel_comp = unrel_comp.average().filter(l_freq=0, h_freq=40)

    # Make anatomical model
    subject = 'fsaverage'
    model = mne.make_bem_model(subject=subject, ico=4, conductivity=conductivity, subjects_dir=subjects_dir)
    bem = mne.make_bem_solution(model)
    src = mne.setup_source_space(subject, spacing="oct6", add_dist="patch", subjects_dir=subjects_dir)

    #COMPREHENSION
    # Compute the noise covariance using the baseline interval
    noise_cov_comp = mne.compute_covariance(epochs_comp, tmin=baseline_start, tmax=baseline_end,
                                            method=["shrunk", "empirical"], rank=None, verbose=True) # list of methods means it will pick the best estimator based on log-likelihood and cross-validation (automated model selection in covariance, Engemann and Gramfort)

    raw_fname_comp = directory + sub + '_comp_preproc.fif'
    trans = directory + sub + '-trans.fif'

    # Compute forward solution
    fwd_comp = mne.make_forward_solution(
        raw_fname_comp,
        trans=trans,
        src=src,
        bem=bem,
        meg=True,
        eeg=False,
        mindist=5.0, # I don't know why this is 5 mm, this parameter is the minimum distance of sources from inner skull surface
        ignore_ref=True # Necessary for our system?
    )
    lambda2 = 1.0 / snr ** 2 # Don't know why

    # Calculate inverse operators
    inverse_ident_comp = make_inverse_operator(ident_comp.info, fwd_comp, noise_cov_comp, loose=0.2, depth=0.8)
    inverse_unrel_comp = make_inverse_operator(unrel_comp.info, fwd_comp, noise_cov_comp, loose=0.2, depth=0.8)

    # Calculate STCs using evoked data, inverse operators from above
    stc_ident_comp = apply_inverse(ident_comp, inverse_ident_comp, lambda2, method=method,pick_ori=None)
    stc_unrel_comp = apply_inverse(unrel_comp, inverse_unrel_comp, lambda2, method=method,pick_ori=None)

    # Save STCs
    fname_ident_comp = directory + sub + '_ident_comp'
    stc_ident_comp.save(fname_ident_comp, ftype='stc', overwrite=True)
    fname_unrel_comp = directory + sub + '_unrel_comp'
    stc_unrel_comp.save(fname_unrel_comp, ftype='stc', overwrite=True)

    del raw_fname_comp, fwd_comp, epoch_fname_comp, epochs_comp, noise_cov_comp, inverse_ident_comp, inverse_unrel_comp, stc_ident_comp, stc_unrel_comp

    #PRODUCTION
    # Compute the noise covariance using the baseline interval
    noise_cov_prod = mne.compute_covariance(epochs_prod, tmin=baseline_start, tmax=baseline_end, method=["shrunk", "empirical"], rank=None, verbose=True)

    raw_fname_prod = directory + sub + '_prod_preproc.fif'

    # Compute forward solution
    fwd_prod = mne.make_forward_solution(
        raw_fname_prod,
        trans=trans,
        src=src,
        bem=bem,
        meg=True,
        eeg=False,
        mindist=5.0,
        ignore_ref=True
    )
    lambda2 = 1.0 / snr ** 2

    # Calculate inverse operators
    inverse_ident_prod = make_inverse_operator(ident_prod.info, fwd_prod, noise_cov_prod, loose=0.2, depth=0.8)
    inverse_unrel_prod = make_inverse_operator(unrel_prod.info, fwd_prod, noise_cov_prod, loose=0.2, depth=0.8)

    # Calculate STCs using evoked data, inverse operators from above
    stc_ident_prod = apply_inverse(ident_prod, inverse_ident_prod, lambda2, method=method, pick_ori=None)
    stc_unrel_prod = apply_inverse(unrel_prod, inverse_unrel_prod, lambda2, method=method, pick_ori=None)

    # Save STCs
    fname_ident_prod = directory + sub + '_ident_prod'
    stc_ident_prod.save(fname_ident_prod, ftype='stc', overwrite = True)
    fname_unrel_prod = directory + sub + '_unrel_prod'
    stc_unrel_prod.save(fname_unrel_prod, ftype='stc', overwrite = True)


if __name__ == "__main__":
    # parser = argparse.ArgumentParser(description="Obtain STC file for a subject.")
    # parser.add_argument("sub", type=str, help="Subject number (e.g., 101)")
    #
    # args = parser.parse_args()
    # source_localization(args.sub)

    for sub in ['R3250', 'R3254', 'R3261', 'R3264', 'R3270','R3271','R3272','R3273','R3275','R3277','R3279','R3285','R3286','R3289','R3290']:
        source_localization(sub)